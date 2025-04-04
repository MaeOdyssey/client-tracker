# therapist_tracker/utils/restore_from_excel.py

import os
import json
from openpyxl import load_workbook
from tkinter import filedialog
from utils.path_helper import get_resource_path


DATA_FILE = get_resource_path("data/clients_data.json")

def restore_from_excel():
    # 🗂️ Prompt user to choose which backup to load
    file_path = filedialog.askopenfilename(
        title="Select Backup File",
        filetypes=[("Excel Files", "*.xlsx")],
        initialdir=get_resource_path("backup")
    )

    if not file_path:
        raise FileNotFoundError("No backup file selected.")

    if not os.path.exists(file_path):
        raise FileNotFoundError("Selected backup file does not exist.")

    wb = load_workbook(file_path)
    ws = wb.active

    restored_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header row
        name, date, time, fee, client_paid, insurance_paid, _ = row

        try:
            fee = float(str(fee).replace("$", "").replace(",", "").strip())
            client_paid = float(str(client_paid).replace("$", "").replace(",", "").strip())
            insurance_paid = float(str(insurance_paid).replace("$", "").replace(",", "").strip())
        except ValueError:
            continue  # Skip invalid rows

        if name not in restored_data:
            restored_data[name] = {
                "name": name,
                "client_id": f"restored-{len(restored_data)}",
                "sessions": []
            }

        restored_data[name]["sessions"].append({
            "date": date,
            "time": time,
            "fee": fee,
            "client_paid": client_paid,
            "insurance_paid": insurance_paid
        })

    # Convert to list format
    final_data = list(restored_data.values())

    # Save to clients_data.json
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    with open(DATA_FILE, "w") as f:
        json.dump(final_data, f, indent=4)

    return DATA_FILE
