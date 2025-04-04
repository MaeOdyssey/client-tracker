import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import uuid
from utils.path_helper import get_resource_path
from utils.payments_helper import load_payments

DATA_FILE = get_resource_path("data/clients_data.json")
PAYMENTS_FILE = get_resource_path("data/payments.json")

PAID_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # soft green
UNPAID_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # soft red

def export_to_excel():
    if not os.path.exists(DATA_FILE):
        raise FileNotFoundError("Client data file not found.")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    uid = uuid.uuid4().hex[:5].upper()
    export_file = f"backup/theracord_backup_{timestamp}_{uid}.xlsx"

    with open(DATA_FILE, "r") as f:
        clients = json.load(f)

    payments = load_payments()

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Sessions"

    headers = [
        "Client Name", "Email", "Phone",
        "Session Date", "Session Time",
        "Fee Charged", "Paid by Client", "Paid by Insurance", "Remaining Balance"
    ]
    ws.append(headers)

    for client in clients:
        name = client.get("name", "Unknown")
        email = client.get("email", "")
        phone = client.get("phone", "")
        for session in client.get("sessions", []):
            session_id = session.get("session_id")
            fee = session.get("fee", 0.0)
            payment = next((p for p in payments if p["session_id"] == session_id), {})
            client_paid = payment.get("client_paid", 0.0)
            insurance_paid = payment.get("insurance_paid", 0.0)
            total_paid = client_paid + insurance_paid
            remaining = fee - total_paid

            row_data = [
                name, email, phone,
                session.get("date", ""),
                session.get("time", ""),
                f"${fee:.2f}",
                f"${client_paid:.2f}",
                f"${insurance_paid:.2f}",
                f"${remaining:.2f}"
            ]

            row = ws.max_row + 1
            ws.append(row_data)

            fill = PAID_FILL if remaining <= 0 else UNPAID_FILL
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill

    os.makedirs(os.path.dirname(export_file), exist_ok=True)
    wb.save(export_file)
    wb.close()
    del wb

    return export_file
